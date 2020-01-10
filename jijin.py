#!/usr/bin/python
#coding:utf-8
__version__ = 20200105



from PIL import Image
import pytesseract,math,os ,csv,codecs
from prettytable import PrettyTable 
from pathlib import PureWindowsPath as pp
import time
import openpyxl
from pprint import pprint


from mystr import splitall,batchremovestr


'''
1. iPhone wx screen shot
2. Picsew 长截图
3. Tesseract
'''


ocr = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
pytesseract.pytesseract.tesseract_cmd =ocr

path = r'M:\MyProject\JiJin'
threshold = 189
xls = r'M:\MyProject\JiJin\t.xlsx'
# xls = r'N:\Doc\财务统计.xlsx'

# (left, upper, right, lower)
# The right can also be represented as (left+width)
# and lower can be represented as (upper+height)

# jump = 230
#(750, 1334)
# print(im.size)


def modificate(text):
    rlist = ['，','预约中','”',']',',']
    return batchremovestr(rlist,text)


def formatnumber(number):
    n = number.split('.')
    if len(n) >= 2:
        number = n[0]+n[1]+'.'+n[-1]
    return number

def initTable(threshold=150):
    '''140'''
    table = []
    for i in range(256):
        if i < threshold:
            table.append(0)
        else:
            table.append(1)
    return table

    
def formattable_jj(jdict):
    t = PrettyTable()
    t.field_names = ['基金','利率','天数','Status']
    t.align['基金'] = 'l'
    t.sortby = '基金'
    for k in jdict.keys():
        t.add_row([k,jdict[k][0],jdict[k][1],jdict[k][2]])
    print(t)


def formattable_mm(mdict):
    m = PrettyTable()
    m.field_names = ['基金','数量','Status']
    m.align['基金'] = 'l'
    m.sortby = '基金'
    for k in mdict.keys():
        m.add_row([k,mdict[k][0],mdict[k][1]])
    print(m)


def readjj(path,threshold,height=142):
    '''Crop the image every 245 pix'''
    seplist = ['\n',' ',]
    jdict = {}
    n = 0
    img = os.path.join(path,'j.PNG')
    im = Image.open(img).convert('L') 
    while True:
        try:    
            box = (0,height*n,400,height*(n+1))
            ni = im.crop(box)
            # ni.show()
            ni = ni.point(initTable(threshold),'1')
            content = pytesseract.image_to_string(ni,lang='chi_sim')
            content = splitall(seplist,modificate(content))
            # print('='*40)
            # print(content)
            name = content[0]
            for z in range(1,len(content)):
                if '%' in content[z]:
                    rate = content[z].split('%')[0][:4]
                    break
                else:
                    rate = 0
            if rate == 0:  
                for z in range(1,len(content)):                   
                    if '.' in content[z]: 
                        rate = content[z].split('%')[0][:4]
                        break
            for z in range(2,len(content)):
                if '最快当天' in content[z]:
                    days = 1
                    break
                elif '天' in content[z]: 
                    days = content[z].split('天')[0]
                    break
                elif '月' in content[z]: 
                    days = content[z].split('月')[0][0]
                    days = str(int(days)*30)
                    break
                else:
                    days = 1
            jdict[name] = [rate,days,' ']
            if content == []: break
            n += 1
        except IndexError:
            break
    # pprint(jdict) # details
    return jdict


def readsaving(path,threshold,height=157):
    '''Crop the image every 270 pix'''
    seplist = ['\n',' ',]
    n = 0
    jdict = {}
    img = os.path.join(path,'m.PNG')
    im = Image.open(img).convert('L') 
    while True:
        try:    
            box = (0,height*n,400,height*(n+1))
            ni = im.crop(box)
            # ni.show()
            ni = ni.point(initTable(threshold),'1')
            content = pytesseract.image_to_string(ni,lang='chi_sim')
            content = splitall(seplist,modificate(content))
            # print(content)
            name = content[0]
            for z in range(1,len(content)):
                if '.' in content[z]:  
                    amount = formatnumber(content[z])
                    break   
            # print('='*40)
            jdict[name] = [amount,' ']
            if content == []: break
            n += 1
        except IndexError:
            break
    return jdict


def toexcel(xls,jdict,mdict):
    '''Output to excel'''
    wb = openpyxl.load_workbook(xls)
    sheet = wb['理财明细']
    for r in range(2,56):   
        sheet.cell(row=r,column=4).value = None 
        sheet.cell(row=r,column=6).value = None
        fundname = sheet.cell(row=r,column=2).value
        # print(fundname)
        try:
            sheet.cell(row=r,column=4).value = float(jdict[fundname][0])
            jdict[fundname][2] = 'match'
            sheet.cell(row=r,column=6).value = float(mdict[fundname][0])
            mdict[fundname][1] = 'match'
        except KeyError:
            pass
    try:
        wb.save(xls)
        os.startfile(xls)
    except PermissionError:
        raise WindowsError('FileOpen')


def main():
    # Update finance product
    jdict = readjj(path,190)
    # pprint(jdict)
    # formattable_jj(jdict)

    # # Update money repository
    mdict = readsaving(path,190)
    # # pprint(mdict)
    # formattable_mm(mdict)

    toexcel(xls,jdict,mdict)
    formattable_jj(jdict)
    formattable_mm(mdict)



if __name__ == "__main__":
    main()
    
