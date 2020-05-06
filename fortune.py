#!/usr/bin/python
#coding:utf-8

__version__ = 20200224


from PIL import Image
import math,os ,csv,codecs
from prettytable import PrettyTable 
# from pathlib import PureWindowsPath as pp
import time
import openpyxl
from pprint import pprint

from mystr import batchremovestr,batchreplacestr
from myocr import init_client,ocr_baidu

'''
1. iPhone wx screen shot
2. Picsew 长截图
3. wx 转原图
4. Baidu OCR
'''


path = r'M:\MyProject\JiJin'
# xls = r'M:\MyProject\JiJin\t.xlsx'
xls = r'N:\Doc\财务统计.xlsx'


class jijin():
    def __init__(self,name):
        self.name = name
        self.rate = None
        self.days = None
        self.amount = None
        self.match = None


def get_file_content(filePath):
    with open(filePath, 'rb') as fp:
        return fp.read()

def find_txt(txt):
    for y in ['持有','收益','短期波动','不改','有元']: #过滤说明文字
        if y in txt:
            return True
    if txt[0] in ['+','-','0']: #过滤收益
        return True
    return False

def modificate(text):
    rlist = ['假期收益已提前发放','预约中',',','支持','零钱通','买入','净值型',
        '智能','NEW','灵活申赎','随买随取最快当到账','HOT',
        # '产品家','产品读','短期涨跌不改长期配置价值',
        '职','期收已发放','已发'
    ]
    return batchremovestr(rlist,text)

def repl(text):
    tlist = [('室盈','富盈')]
    return batchreplacestr(tlist,text)
  
def formattable_jj(jdict):
    t = PrettyTable()
    t.field_names = ['基金','利率','天数','数量','Match']
    t.align['基金'] = 'l'
    t.sortby = '基金'
    for k in jdict.keys():
        jj = jdict[k]
        t.add_row([ jj.name,jj.rate,jj.days,jj.amount,jj.match ])
    print(t)


def readsaving(ocrclient,jdict,path):
    '''Crop the image every 271 pix'''
    img = os.path.join(path,'s.png')
    im = Image.open(img).convert('L') 
    im.save('tmp.png', 'PNG')
    data = get_file_content('tmp.png')
    # data = get_file_content(img)
    try:    
        txtlist = ocr_baidu(ocrclient,data)
        # pprint(txtlist)
        newlist = []
        for x in txtlist:
            if not find_txt(x):
                newlist.append(x)  
        # pprint(newlist)
        mark = False
        j = {}
        for x in newlist:
            try:
                int(x[0])
                amount = x.split('+')[0]
                if mark:
                    j[mark] = amount.replace(',','')
                    mark = False
            except ValueError:
                name = x.split('(')[0]
                name = modificate(repl(name))
                j[name] = ''
                mark = name
        for x,y in j.items():
            j = jijin(x)
            j.amount = y
            jdict[x] = j
    except:
        raise
    # print(jdict) # details
    return jdict


def readjj(ocrclient,jdict,path,height=243.6):
    '''Crop the image every 271 pix'''
    n = 0
    img = os.path.join(path,'j.png')
    im = Image.open(img).convert('L') 
    while True:
        try:    
            box = (0,height*n,800,height*(n+1))
            ni = im.crop(box)
            # ni.show()
            ni.save('tmp.png', 'PNG')
            data = get_file_content('tmp.png')
            txtlist = ocr_baidu(ocrclient,data)
            # pprint(txtlist)
            name = modificate(txtlist[0])
            rateday = txtlist[1].split('%')
            rate = rateday[0]
            if '天' in rateday[1]:
                days = rateday[1].replace('天','')
            elif '个月' in rateday[1]:
                days = rateday[1].replace('个月','')
            else:
                days = txtlist[2].replace('天','')
            if modificate(days) == '':
                days = 1
            if name not in jdict.keys():
                j = jijin(name)
                jdict[name] = j
            jdict[name].rate = rate
            jdict[name].days = days
            if txtlist == []:
                break
            n += 1
            time.sleep(1)
        except IndexError:
            break
    # pprint(jdict) # details
    return jdict


def toexcel(xls,jdict):
    '''Output to excel'''
    wb = openpyxl.load_workbook(xls)
    sheet = wb['理财明细']
    for r in range(2,56):   
        sheet.cell(row=r,column=4).value = None  # rate
        sheet.cell(row=r,column=6).value = None  # amount
        fundname = sheet.cell(row=r,column=2).value
        # print(fundname)
        if fundname in jdict.keys():
            try:           
                sheet.cell(row=r,column=4).value = float(jdict[fundname].rate) 
            except:
                pass               
            try:                
                sheet.cell(row=r,column=6).value = float(jdict[fundname].amount)
            except:
                pass
            jdict[fundname].match = 'Yes'
    try:
        wb.save(xls)
        os.startfile(xls)
    except PermissionError:
        raise WindowsError('FileOpen')


def main():
    ''' Update finance product'''
    ocr = init_client()
    jdict = {}
    jdict = readsaving(ocr,jdict,path)
    jdict = readjj(ocr,jdict,path)
    # formattable_jj(jdict)
    toexcel(xls,jdict)
    formattable_jj(jdict)



if __name__ == "__main__":
    main()
    
