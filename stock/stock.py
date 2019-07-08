#!/usr/bin/python3
#coding:utf-8
# test in Win

import openpyxl, argparse, time,os
import pyautogui as auto

xls = r'E:\UT\stock.xlsx'
wp = os.path.dirname(os.path.realpath(__file__))
wt = r'D:\Profile\Desktop\wt.txt'


# xl = r'C:\D\JG\我的坚果云\newstock.xlsx'
# s = r'C:\D\JG\我的坚果云\wt.txt'

#印花税：单向收取，卖出成交金额的千分之一（1‰）。 
#过户费：按成交股票的金额×0.02‰收取，单位：元。双向收取（仅上海股票收取）。
#手续费： 5元 
#对账单

def readtransaction_sh(s):
# 发生日期 业务名称 证券代码 证券名称 成交均价 成交数量 成交金额 股份余额 手续费 印花税 过户费 其他费 发生金额 资金余额 委托编号 委托价格 委托数量 股东代码 资金帐号 币种 备注 
    with open(s,'r') as f:
        transactions = {}
        n = 0
        for x in f.readlines()[3:]:
            e = x.split() 
            if e[1] in ['证券买入清算','证券卖出清算']:
                x = -1 if e[1] == '证券买入清算' else 1
                ttime = e[0]
                stock = e[3]
                ty = '买' if e[1] == '证券买入清算' else '卖'
                price = round(float(e[4]),2)
                qty = float(e[5])
                final = round(float(e[12]),2)
                v = [ttime,stock,ty,price,x*qty,final]
                # print(v)
                transactions[n] = v
                n += 1
    # print(t)
    return transactions


def readtransaction_hk(s):
#交易市场/发生日期/发生时间/业务名称/证券代码/证券名称/成交价格(港币)/成交数量/成交金额(人民币)/股份余额/手续费(人民币)/印花税(人民币)/交易征费(人民币)/股份交收费(人民币)/交易费(人民币)/系统使用费(人民币)/发生金额(人民币)/资金余额(人民币)/结算汇率/委托编号/委托价格(港币)/委托数量/股东代码/资金帐号/币种
    with open(s,'r') as f:
        transactions = {}
        n = 0
        for x in f.readlines()[3:]:
            e = x.split() 
            if e[3] in ['证券买入清算','证券卖出清算']:
                x = -1 if e[3] == '证券买入清算' else 1
                ttime = e[1]
                stock = e[5]
                ty = '买' if e[3] == '证券买入清算' else '卖'
                price = round(float(e[6]),2)
                qty = float(e[7])
                final = round(float(e[16]),2)
                v = [ttime,stock,ty,price,x*qty,final]
                # print(v)
                transactions[n] = v
                n += 1
    # print(t)
    return transactions


def addtransaction(xls,transactions,page):
    wb = openpyxl.load_workbook(xls)
    sheet = wb[page]
    max = sheet.max_row
    for x in range(len(transactions)):
        # print(t[i])                    
        for n in range(6) :
            # print(t[i][n])
            sheet.cell( row=(max+x+1),column=(n+1) ).value = transactions[x][n] 
    wb.save(xls)


def readtransaction_g(s):
#序号/交易日期/交易时间/交易场所/交易类型/交易方式/钞汇标志/交易品种/成交价格/单位/成交数量/单位/成交金额/单位
    with open(s,'r') as f:
        transactions = {}
        n = 0
        for x in f.readlines()[7:]:
            # print(x)
            if x.split() != []:
                e = x.split() 
                # print(e)
                if e[4] in ['卖出平仓','买入开仓']:
                    x = -1 if e[4] == '买入开仓' else 1
                    ttime = e[1]
                    gtype = e[7]
                    ty = '买' if e[4] == '买入开仓' else '卖'
                    price = round(float(e[8]),2)
                    qty = float(e[10])
                    final = round(float(e[12]),2)
                    v = [ttime,gtype,ty,price,x*qty,x*final]
                    # print(v)
                    transactions[n] = v
                    n += 1
    # print(transactions)
    return transactions


def addtransaction_g(xls,transactions,page):
    wb = openpyxl.load_workbook(xls)
    sheet = wb[page]
    max = sheet.max_row
    r = 0  
    for x in range(len(transactions)-1,-1,-1):
        # print(t[i])      
        # print(transactions[x])                 
        for n in range(6) :
            # print(t[i][n])
            sheet.cell( row=(max+r+1),column=(n+1) ).value = transactions[x][n] 
        r += 1
    wb.save(xls)


def capture(wp,img,trys=3,confidence=0.9):    
    pic = os.path.join(wp,img+'.png')
    # print(pic)
    trytime = 1
    while trytime < trys:
        try:
            button = auto.locateCenterOnScreen(pic,grayscale=True,confidence=confidence)
            time.sleep(1)
            break            
        except TypeError:
            time.sleep(15)
            trytime += 1
            print(f'try to locate {img} {str(trytime)}')
            continue
        except OSError as e:
            print(e)
            return e
    else:
        print(f"Max tries reach, not able to locate option {img}")
        button = None
        # return f"Max tries reach, not able to locate option {img}"
        # raise(f'Find no {img}')
    return button


def click_flow():
    flow = ['query',
            'output',
            'tofile',
            'browse',
            'save',
            'yes']
    try:
        os.remove(wt)
    except:
        pass
    for x in flow:
        findbutton(x)
        # button = capture(wp,x)
        # if isinstance(button,tuple):
        #     auto.click(button)
        #     # auto.click(button)
        # else:
        #     return button
    time.sleep(2)
    auto.keyDown('alt')
    auto.keyDown('f4')
    auto.keyUp('alt')
    auto.keyUp('f4')


def findbutton(img,**args):
    button = capture(wp,img,**args)
    if isinstance(button,tuple):
        auto.click(button)
        return True
    else:
        return False


def savequery():
    print('统计股票')
    #沪深
    if findbutton('transactionY') is False:
        findbutton('transaction')
    click_flow()
    transactions = readtransaction_sh(wt)
    addtransaction(xls,transactions,'trans')    
    #港股
    time.sleep(1)
    findbutton('hk')     
    if findbutton('hktranY') is False:
        findbutton('hktransaction')
    click_flow()
    transactions = readtransaction_hk(wt)
    addtransaction(xls,transactions,'港股')
    os.startfile(xls)


def main():
    parser = argparse.ArgumentParser(description = 'Stock tool')
    group = parser.add_mutually_exclusive_group()
    group.add_argument('-s',action="store_true", help='Shanghai Stock')
    group.add_argument('-h',action="store_true", help='Help')
    group.add_argument('-g',action="store_true", help='Golden')

    args = parser.parse_args()

    if args.h:
        parser.print_help()
    elif args.g:
        transactions = readtransaction_g(wt)
        addtransaction_g(xls,transactions,'gold')        
    else:
        savequery()

if __name__ == "__main__":   
    try:
        savequery()
    except PermissionError as e:
        print(e)
        print('Is file being opened?')