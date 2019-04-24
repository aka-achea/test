#!/usr/bin/python
#coding:utf-8
#Python3
# test in Win

import openpyxl

s = r'D:\Profile\Desktop\wt.txt'
xl = r'E:\UT\stock.xlsx'

# xl = r'C:\D\JG\我的坚果云\newstock.xlsx'
# s = r'C:\D\JG\我的坚果云\wt.txt'

#印花税：单向收取，卖出成交金额的千分之一（1‰）。 
#过户费：按成交股票的金额×0.02‰收取，单位：元。双向收取（仅上海股票收取）。
#手续费： 5元 
#对账单

def readtransaction(s):
    with open(s,'r') as f:
        t = {}
        n = 0
        for i in f.readlines()[3:]:
            e = i.split() 
            if e[1] in ['证券买入清算','证券卖出清算']:
                x = -1 if e[1] == '证券买入清算' else 1
                ttime = e[0]
                stock = e[3]
                ty = '买' if e[1] == '证券买入清算' else '卖'
                price = round(float(e[4]),2)
                qty = float(e[5])
                final = round(float(e[12]),2)
                v = [ttime,stock,ty,price,x*qty,final]
                # print(v)
                t[n] = v
                n += 1
    # print(t)
    return t

def stock(xl,t):
    wb = openpyxl.load_workbook(xl)
    sheet = wb['trans']
    max = sheet.max_row
    for i in range(len(t)):
        # print(t[i])                    
        for n in range(6) :
            # print(t[i][n])
            sheet.cell( row=(max+i+1),column=(n+1) ).value = t[i][n] 
    wb.save(xl)

try:
    t = readtransaction(s)
    stock(xl,t)
except PermissionError as e:
    print(e)
    print('Is file being opened?')