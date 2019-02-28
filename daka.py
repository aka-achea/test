#!/usr/bin/python3
#coding:utf-8

# tested in Win

import openpyxl,math,re,os,configparser
import PIL
from PIL import ImageFont, Image, ImageDraw

#customized module
from mylog import mylogger,get_funcname

confile = 'daka.ini'
scriptpath = os.path.dirname(os.path.realpath(__file__))
confilepath = os.path.join(scriptpath,confile)
config = configparser.ConfigParser()
config.read(confilepath)
workpath = config['setting']['workpath']
dakarecord = config['setting']['daka']
walkrecord = config['setting']['walk']
logfile = os.path.join(workpath,config['setting']['log'])
logfilelevel = 10
output = os.path.join(workpath,config['setting']['output'])
if os.path.exists(output):
    os.remove(output)

############ 图片设置 ############
fontfile = config['picture']['fontfile']
background = config['picture']['background']
picsum = config['picture']['picsum']
picon = config['picture']['picon']
picoff = config['picture']['picoff']
size = int(config['picture']['size'])
color = config['picture']['color']
c = color.split(',')
color = ( int(c[0]),int(c[1]),int(c[2]) )
fontfilepath = os.path.join(workpath,fontfile)
font = ImageFont.truetype(fontfile,size,encoding='utf-8')
imageFile = os.path.join(workpath,background)

############ 统计考勤 ############

def tj(dakarecord):
    l = mylogger(logfile,logfilelevel,get_funcname()) 
    wb = openpyxl.load_workbook(os.path.join(workpath,dakarecord))
    tjdic ={}
    sheet1 = wb['原始记录']  
    for i in range(4,sheet1.max_row+1):
        status = sheet1.cell(row=i,column=9).value  #打卡状态
        if  re.search('打卡无效',status):
            l.debug('Row '+str(i)+' 打卡无效')
            continue
        name = sheet1.cell(row=i,column=1).value  # 姓名
        if name not in tjdic: 
            tjdic[name] = {}
        dk = sheet1.cell(row=i,column=8).value # 打卡日期时间
        day = dk.split(' ')[0].split('-')[-1].replace('0','') 
        l.debug(name+' '+day+' '+dk)
        if day not in tjdic[name]:
            tjdic[name][day] = {}
        ti = dk.split(' ')[-1] # 打卡时间
        h = int(ti.split(':')[0])
        m = int(ti.split(':')[1])
        hm = h+m/60
        if hm > 12: # 下班
            tjdic[name][day]['off'] = hm
            tjdic[name][day]['rawoff'] = ti
        else:   # 上班
            tjdic[name][day]['on'] = hm
            tjdic[name][day]['rawon'] = ti
    l.debug(tjdic)
    return tjdic

def fillempty(tjdic):
    l = mylogger(logfile,logfilelevel,get_funcname()) 
    for i in tjdic:
        name = i
        for d in tjdic[i]:
            day = d
            if 'on' not in tjdic[i][d]:
                tjdic[name][day]['on'] = 9
                tjdic[name][day]['rawon'] = '9:00'
                l.debug(name+' '+day+' 9:00 未打卡')
            if 'off' not in tjdic[i][d]:
                tjdic[name][day]['off'] = 18
                tjdic[name][day]['rawoff'] = '18:00'
                l.debug(name+' '+day+' 18:00 未打卡')
    l.debug(tjdic)     
    return tjdic

def cal(tjdic):
    l = mylogger(logfile,logfilelevel,get_funcname()) 
    result = {}
    for i in tjdic:
        name = i
        result[name] = {}    
        maxt = 0
        avgt = 0
        sumt = 0
        maxoff = ''
        for d in tjdic[i]:
            day = d
            off = tjdic[name][day]['off']
            rawoff = tjdic[name][day]['rawoff']
            on = tjdic[name][day]['on']
            sumt = sumt + off - on
            avgt = avgt + on
            if off > maxt: 
                maxt = off
                maxoff = rawoff
        avgt = avgt / len(tjdic[i])
        l.debug(name+' '+str(maxt)+' '+str(avgt)+' '+str(sumt))
        result[name]['夜归人榜'] = maxoff
        result[name]['平均早鸟榜'] = avgt
        result[name]['办公室达人榜'] = sumt
    l.debug(result)
    return result

def str2hm(i): # convert string to Hour:Minute
    h = int(i)
    h = '0'+str(h) if h <10 else str(h)
    m = round((i - int(i))*60)
    m = '0'+str(m) if m <10 else str(m)
    hm = h+':'+m
    return hm

def winner(result,bang): # sort according to score
    l = mylogger(logfile,logfilelevel,get_funcname()) 
    with open(output,'a') as f:
        l.debug(bang)
        f.writelines('='*5+bang+'='*5+'\n')
        scorenamedic = {}
        count = 0
        for i in result:
            try:
                scorenamedic[result[i][bang]] = i    
            except:
                l.error(bang+' '+i+' 有问题' )
        winlist = {}
        for i in sorted(scorenamedic,reverse=(True if bang != '平均早鸟榜' else False ) ):
            if bang == '办公室达人榜':
                hm = str2hm(i)
                hm = hm.split(':')[0]+'小时'+hm.split(':')[1]+'分钟'
            elif bang == '平均早鸟榜':
                hm = str2hm(i)
            else:
                hm = i
            count += 1
            l.debug('%s %s %s' % (count , scorenamedic[i],hm))
            # f.writelines('%s %s %s \n' % (count , scorenamedic[i],hm))
            winlist[count] = (scorenamedic[i],hm)
        l.debug(winlist)
        toplist = buildtoplist(winlist)
        for t in toplist:
            f.writelines('%s %s %s \n' % (t[0],t[1],t[2]))
    return toplist

def buildtoplist(winlist):
    l = mylogger(logfile,logfilelevel,get_funcname()) 
    toplist = []
    lastrank = 1
    lastvalue = ''
    for t in range(len(winlist)):
        rank = t+1
        if rank > 10: break
        name = winlist[rank][0]
        value = str(winlist[rank][1])
        if value == lastvalue:
            rank = lastrank
        l.debug([rank,name,value])
        toplist.append([str(rank),name,value])
        lastvalue = value
        lastrank = rank
    l.debug(toplist)
    return toplist

# wlist = [[position,name,value]...]
def windraw(wlist,imout,title):
    l = mylogger(logfile,logfilelevel,get_funcname()) 
    im = Image.open(imageFile)
    draw = ImageDraw.Draw(im)
    x,y = (70,50) #初始左上角的坐标
    xstep = 100
    ystep = 40
    texttile = title[0]
    draw.text( (x+xstep*2,y),texttile,color,font=font )
    x,y = (x, y+ystep*2)
    draw.text( (x,y),title[1],color,font=font )
    draw.text( (x+xstep*2,y),title[2],color,font=font )
    draw.text( (x+xstep*4,y),title[3],color,font=font )
    x,y = (x, y+ystep)
    for i in range(len(wlist)):
        textposition = wlist[i][0]
        textname = wlist[i][1]
        textvalue = str(wlist[i][2])
        l.debug(textposition+' '+textname+' '+str(textvalue))
        draw.text( (x,y),textposition,color,font=font )
        draw.text( (x+xstep*2,y),textname,color,font=font )
        draw.text( (x+xstep*4,y),textvalue,color,font=font )
        x,y = (x,y+ystep)
    draw = ImageDraw.Draw(im)
    im.save(imout)


############ 统计步数 ############
def walk(walkrecord,result): # add to result
    l = mylogger(logfile,logfilelevel,get_funcname()) 
    wb = openpyxl.load_workbook(os.path.join(workpath,walkrecord))
    sheet = wb['钉钉运动数据导出1']  
    for i in range(3,sheet.max_row+1):
        name = sheet.cell(row=i,column=3).value  # 姓名
        if name not in result:
            result[name] = {} 
        if '健步狂人榜' not in result[name]: 
            # namestepdic[name] = {}
            tstep = 0
        steps = sheet.cell(row=i,column=5).value # 打卡日期时间
        l.debug(name+' '+str(steps))
        tstep = tstep + steps
        result[name]['健步狂人榜'] = tstep
    l.debug(result)
    return result


def main():
    l = mylogger(logfile,logfilelevel,get_funcname()) 
    try:        
        tjdic = tj(dakarecord)
        tjdic = fillempty(tjdic)
        result = cal(tjdic)
        winonlist = winner(result,'平均早鸟榜')
        winofflist = winner(result,'夜归人榜')
        winsumlist = winner(result,'办公室达人榜')
        result = walk(walkrecord,result)
        winsteplist = winner(result,'健步狂人榜')

        piclist = [['办公室达人榜','排名','名字','工作时长（小时）',winsumlist],
                    ['平均早鸟榜','排名','名字','平均打卡时间',winonlist],
                    ['夜归人榜','排名','名字','最晚打卡时间',winofflist],
                    ['健步狂人榜','排名','名字','健身步数',winsteplist] ]

        for b in piclist: 
            outpath = os.path.join(workpath,b[0]+'.jpg')
            title = [b[0],b[1],b[2],b[3]]
            l.debug(outpath)
            windraw(b[4],outpath,title)

        # title = ['办公室达人','排名','名字','工作时长（小时）']
        # outpath = os.path.join(workpath,picsum)
        # windraw(winsumlist,outpath,title)

        # title = ['早鸟榜','排名','名字','平均打卡时间']
        # outpath = os.path.join(workpath,picon)
        # windraw(winonlist,outpath,title)

        # title = ['夜归人','排名','名字','最晚打卡时间']
        # outpath = os.path.join(workpath,picoff)
        # windraw(winofflist,outpath,title)

    except PermissionError as e:
        print(e)
        print('Is file being opened?')



if __name__=='__main__':
    main()


